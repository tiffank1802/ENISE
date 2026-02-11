"""
Script pour générer un fichier Excel avec les hypothèses et calculs DCF
pour l'acquisition de Bricorama France.
Toutes les cellules de calcul contiennent des formules Excel (pas de valeurs en dur).
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter

# Créer le workbook
wb = Workbook()

# ==============================================================================
# FEUILLE 1 : HYPOTHESES
# ==============================================================================
ws_hyp = wb.active
ws_hyp.title = "Hypotheses"

# Style pour les titres
title_font = Font(bold=True, size=14, color="FFFFFF")
header_font = Font(bold=True, size=11)
title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
light_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Titre principal
ws_hyp.merge_cells('A1:D1')
ws_hyp['A1'] = "HYPOTHÈSES - ACQUISITION BRICORAMA FRANCE"
ws_hyp['A1'].font = title_font
ws_hyp['A1'].fill = title_fill
ws_hyp['A1'].alignment = Alignment(horizontal='center')

# Section 1: Données de l'année de base
row = 3
ws_hyp[f'A{row}'] = "1. DONNÉES DE L'ANNÉE DE BASE (N = 2024)"
ws_hyp[f'A{row}'].font = header_font
ws_hyp[f'A{row}'].fill = light_fill
ws_hyp.merge_cells(f'A{row}:D{row}')

row = 4
data_base = [
    ("Chiffre d'Affaires (CA_N)", 500000000, "€", "B4"),
    ("EBE (EBE_N)", 40000000, "€", "B5"),
    ("Résultat d'Exploitation (EBIT_N)", 30000000, "€", "B6"),
    ("BFR (BFR_N)", 50000000, "€", "B7"),
    ("Capex (Capex_N)", 12000000, "€", "B8"),
    ("Dettes financières totales", 80000000, "€", "B9"),
    ("Trésorerie", 10000000, "€", "B10"),
]

for i, (label, value, unit, cell_ref) in enumerate(data_base):
    ws_hyp[f'A{row+i}'] = label
    ws_hyp[f'B{row+i}'] = value
    ws_hyp[f'B{row+i}'].number_format = '#,##0'
    ws_hyp[f'C{row+i}'] = unit
    ws_hyp[f'A{row+i}'].border = thin_border
    ws_hyp[f'B{row+i}'].border = thin_border
    ws_hyp[f'C{row+i}'].border = thin_border

# Section 2: Taux et hypothèses de croissance
row = 12
ws_hyp[f'A{row}'] = "2. TAUX ET HYPOTHÈSES DE CROISSANCE"
ws_hyp[f'A{row}'].font = header_font
ws_hyp[f'A{row}'].fill = light_fill
ws_hyp.merge_cells(f'A{row}:D{row}')

row = 13
taux_data = [
    ("Taux d'impôt sur les sociétés", 0.25, "%", "B13"),
    ("Taux d'actualisation (WACC)", 0.08, "%", "B14"),
    ("Taux de croissance du volume", 0.05, "%", "B15"),
    ("Taux d'inflation prix de vente", 0.05, "%", "B16"),
    ("Taux d'inflation coûts", 0.05, "%", "B17"),
    ("Horizon de projection (années)", 5, "ans", "B18"),
    ("Taux de croissance long terme (g_inf)", 0.02, "%", "B19"),
    ("Multiple EBE pour dette max", 3.5, "x", "B20"),
    ("Taux dette senior", 0.05, "%", "B21"),
]

for i, (label, value, unit, cell_ref) in enumerate(taux_data):
    ws_hyp[f'A{row+i}'] = label
    ws_hyp[f'B{row+i}'] = value
    if unit == "%":
        ws_hyp[f'B{row+i}'].number_format = '0.00%'
    else:
        ws_hyp[f'B{row+i}'].number_format = '#,##0.0'
    ws_hyp[f'C{row+i}'] = unit if unit != "%" else ""
    ws_hyp[f'A{row+i}'].border = thin_border
    ws_hyp[f'B{row+i}'].border = thin_border
    ws_hyp[f'C{row+i}'].border = thin_border

# Section 3: Ratios calculés (avec formules)
row = 23
ws_hyp[f'A{row}'] = "3. RATIOS CALCULÉS (FORMULES)"
ws_hyp[f'A{row}'].font = header_font
ws_hyp[f'A{row}'].fill = light_fill
ws_hyp.merge_cells(f'A{row}:D{row}')

row = 24
ws_hyp['A24'] = "Marge EBE (EBE/CA)"
ws_hyp['B24'] = "=B5/B4"
ws_hyp['B24'].number_format = '0.00%'
ws_hyp['C24'] = "Formule: EBE_N / CA_N"

ws_hyp['A25'] = "Marge EBIT (EBIT/CA)"
ws_hyp['B25'] = "=B6/B4"
ws_hyp['B25'].number_format = '0.00%'
ws_hyp['C25'] = "Formule: EBIT_N / CA_N"

ws_hyp['A26'] = "Ratio Capex/CA"
ws_hyp['B26'] = "=B8/B4"
ws_hyp['B26'].number_format = '0.00%'
ws_hyp['C26'] = "Formule: Capex_N / CA_N"

ws_hyp['A27'] = "Ratio BFR/CA"
ws_hyp['B27'] = "=B7/B4"
ws_hyp['B27'].number_format = '0.00%'
ws_hyp['C27'] = "Formule: BFR_N / CA_N"

ws_hyp['A28'] = "Croissance nominale CA"
ws_hyp['B28'] = "=(1+B15)*(1+B16)-1"
ws_hyp['B28'].number_format = '0.00%'
ws_hyp['C28'] = "Formule: (1+g_vol)*(1+g_prix)-1"

ws_hyp['A29'] = "Dette nette"
ws_hyp['B29'] = "=B9-B10"
ws_hyp['B29'].number_format = '#,##0'
ws_hyp['C29'] = "Formule: Dettes - Trésorerie"

for i in range(24, 30):
    ws_hyp[f'A{i}'].border = thin_border
    ws_hyp[f'B{i}'].border = thin_border
    ws_hyp[f'C{i}'].border = thin_border

# Ajuster largeur des colonnes
ws_hyp.column_dimensions['A'].width = 35
ws_hyp.column_dimensions['B'].width = 18
ws_hyp.column_dimensions['C'].width = 25
ws_hyp.column_dimensions['D'].width = 15

# ==============================================================================
# FEUILLE 2 : CALCULS DCF
# ==============================================================================
ws_dcf = wb.create_sheet("Calculs_DCF")

# Titre
ws_dcf.merge_cells('A1:H1')
ws_dcf['A1'] = "PROJECTION DES FLUX DE TRÉSORERIE (DCF) - 5 ANS"
ws_dcf['A1'].font = title_font
ws_dcf['A1'].fill = title_fill
ws_dcf['A1'].alignment = Alignment(horizontal='center')

# En-têtes
headers = ["Poste", "Année N (Base)", "N+1", "N+2", "N+3", "N+4", "N+5"]
for col, header in enumerate(headers, 1):
    cell = ws_dcf.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.fill = light_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')

# Lignes de calcul avec formules
rows_data = [
    ("Chiffre d'Affaires (CA)", "=Hypotheses!B4", 
     "=B4*(1+Hypotheses!$B$28)", "=C4*(1+Hypotheses!$B$28)", "=D4*(1+Hypotheses!$B$28)", 
     "=E4*(1+Hypotheses!$B$28)", "=F4*(1+Hypotheses!$B$28)"),
    
    ("Croissance CA (%)", "-", 
     "=(C4-B4)/B4", "=(D4-C4)/C4", "=(E4-D4)/D4", "=(F4-E4)/E4", "=(G4-F4)/F4"),
    
    ("EBE (EBITDA)", "=Hypotheses!B5", 
     "=C4*Hypotheses!$B$24", "=D4*Hypotheses!$B$24", "=E4*Hypotheses!$B$24", 
     "=F4*Hypotheses!$B$24", "=G4*Hypotheses!$B$24"),
    
    ("Marge EBE (%)", "=B6/B4", "=C6/C4", "=D6/D4", "=E6/E4", "=F6/F4", "=G6/G4"),
    
    ("Résultat d'Exploitation (EBIT)", "=Hypotheses!B6", 
     "=C4*Hypotheses!$B$25", "=D4*Hypotheses!$B$25", "=E4*Hypotheses!$B$25", 
     "=F4*Hypotheses!$B$25", "=G4*Hypotheses!$B$25"),
    
    ("- Impôt sur Sociétés (25%)", "=B8*Hypotheses!$B$13", 
     "=C8*Hypotheses!$B$13", "=D8*Hypotheses!$B$13", "=E8*Hypotheses!$B$13", 
     "=F8*Hypotheses!$B$13", "=G8*Hypotheses!$B$13"),
    
    ("+ Dotations Amort. (EBITDA-EBIT)", "=B6-B8", 
     "=C6-C8", "=D6-D8", "=E6-E8", "=F6-F8", "=G6-G8"),
    
    ("- Investissements (Capex)", "=Hypotheses!B8", 
     "=C4*Hypotheses!$B$26", "=D4*Hypotheses!$B$26", "=E4*Hypotheses!$B$26", 
     "=F4*Hypotheses!$B$26", "=G4*Hypotheses!$B$26"),
    
    ("BFR (stock)", "=Hypotheses!B7", 
     "=C4*Hypotheses!$B$27", "=D4*Hypotheses!$B$27", "=E4*Hypotheses!$B$27", 
     "=F4*Hypotheses!$B$27", "=G4*Hypotheses!$B$27"),
    
    ("- Variation de BFR", "-", 
     "=C12-B12", "=D12-C12", "=E12-D12", "=F12-E12", "=G12-F12"),
    
    ("", "", "", "", "", "", ""),  # ligne vide
    
    ("= Flux de Trésorerie Disponible (FCF)", "-", 
     "=C6-C9-C13-C11", "=D6-D9-D13-D11", "=E6-E9-E13-E11", 
     "=F6-F9-F13-F11", "=G6-G9-G13-G11"),
]

for row_idx, row_data in enumerate(rows_data, 4):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_dcf.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        if col_idx == 1:
            cell.font = Font(bold=True) if "=" in str(row_data[-1]) and "FCF" in row_data[0] else None
        else:
            if "%" in row_data[0]:
                cell.number_format = '0.00%'
            else:
                cell.number_format = '#,##0'

# Section Actualisation
row = 18
ws_dcf[f'A{row}'] = "ACTUALISATION DES FLUX"
ws_dcf[f'A{row}'].font = header_font
ws_dcf[f'A{row}'].fill = light_fill
ws_dcf.merge_cells(f'A{row}:G{row}')

actual_data = [
    ("Année", "", "1", "2", "3", "4", "5"),
    ("Facteur d'actualisation", "-", 
     "=1/(1+Hypotheses!$B$14)^C19", "=1/(1+Hypotheses!$B$14)^D19", 
     "=1/(1+Hypotheses!$B$14)^E19", "=1/(1+Hypotheses!$B$14)^F19", 
     "=1/(1+Hypotheses!$B$14)^G19"),
    ("FCF Actualisé", "-", 
     "=C15*C20", "=D15*D20", "=E15*E20", "=F15*F20", "=G15*G20"),
]

for row_idx, row_data in enumerate(actual_data, 19):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_dcf.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        if row_idx == 20:
            cell.number_format = '0.0000'
        else:
            cell.number_format = '#,##0'

# Ajuster largeur des colonnes
ws_dcf.column_dimensions['A'].width = 35
for col in range(2, 8):
    ws_dcf.column_dimensions[get_column_letter(col)].width = 15

# ==============================================================================
# FEUILLE 3 : VALORISATION
# ==============================================================================
ws_val = wb.create_sheet("Valorisation")

# Titre
ws_val.merge_cells('A1:D1')
ws_val['A1'] = "VALORISATION DE L'ENTREPRISE"
ws_val['A1'].font = title_font
ws_val['A1'].fill = title_fill
ws_val['A1'].alignment = Alignment(horizontal='center')

# VAN des FCF
row = 3
ws_val[f'A{row}'] = "1. VAN DES FLUX DE TRÉSORERIE"
ws_val[f'A{row}'].font = header_font
ws_val[f'A{row}'].fill = light_fill
ws_val.merge_cells(f'A{row}:D{row}')

ws_val['A4'] = "Somme des FCF actualisés (5 ans)"
ws_val['B4'] = "=SUM(Calculs_DCF!C21:G21)"
ws_val['B4'].number_format = '#,##0'
ws_val['C4'] = "€"
ws_val['D4'] = "Formule: Σ FCF_actualisé"

# Valeur Terminale
row = 6
ws_val[f'A{row}'] = "2. VALEUR TERMINALE"
ws_val[f'A{row}'].font = header_font
ws_val[f'A{row}'].fill = light_fill
ws_val.merge_cells(f'A{row}:D{row}')

ws_val['A7'] = "FCF année 5 (FCF_T)"
ws_val['B7'] = "=Calculs_DCF!G15"
ws_val['B7'].number_format = '#,##0'
ws_val['C7'] = "€"

ws_val['A8'] = "FCF année 6 (FCF_T+1)"
ws_val['B8'] = "=B7*(1+Hypotheses!B19)"
ws_val['B8'].number_format = '#,##0'
ws_val['C8'] = "€"
ws_val['D8'] = "Formule: FCF_T × (1 + g_inf)"

ws_val['A9'] = "Valeur Terminale (non actualisée)"
ws_val['B9'] = "=B8/(Hypotheses!B14-Hypotheses!B19)"
ws_val['B9'].number_format = '#,##0'
ws_val['C9'] = "€"
ws_val['D9'] = "Formule: FCF_T+1 / (WACC - g_inf)"

ws_val['A10'] = "Valeur Terminale Actualisée"
ws_val['B10'] = "=B9/((1+Hypotheses!B14)^Hypotheses!B18)"
ws_val['B10'].number_format = '#,##0'
ws_val['C10'] = "€"
ws_val['D10'] = "Formule: VT / (1+WACC)^5"

# Valeur d'Entreprise
row = 12
ws_val[f'A{row}'] = "3. VALEUR D'ENTREPRISE"
ws_val[f'A{row}'].font = header_font
ws_val[f'A{row}'].fill = title_fill
ws_val[f'A{row}'].font = title_font
ws_val.merge_cells(f'A{row}:D{row}')

ws_val['A13'] = "VALEUR D'ENTREPRISE (VE)"
ws_val['A13'].font = Font(bold=True, size=12)
ws_val['B13'] = "=B4+B10"
ws_val['B13'].number_format = '#,##0'
ws_val['B13'].font = Font(bold=True, size=12)
ws_val['C13'] = "€"
ws_val['D13'] = "Formule: VAN_FCF + VT_actualisée"

# Valeur des Titres
row = 15
ws_val[f'A{row}'] = "4. VALEUR DES TITRES (EQUITY VALUE)"
ws_val[f'A{row}'].font = header_font
ws_val[f'A{row}'].fill = light_fill
ws_val.merge_cells(f'A{row}:D{row}')

ws_val['A16'] = "Dette nette"
ws_val['B16'] = "=Hypotheses!B29"
ws_val['B16'].number_format = '#,##0'
ws_val['C16'] = "€"

ws_val['A17'] = "VALEUR DES TITRES"
ws_val['A17'].font = Font(bold=True, size=12)
ws_val['B17'] = "=B13-B16"
ws_val['B17'].number_format = '#,##0'
ws_val['B17'].font = Font(bold=True, size=12)
ws_val['C17'] = "€"
ws_val['D17'] = "Formule: VE - Dette nette"

# Capacité d'endettement
row = 19
ws_val[f'A{row}'] = "5. CAPACITÉ D'ENDETTEMENT ET APPORT EN CAPITAL"
ws_val[f'A{row}'].font = header_font
ws_val[f'A{row}'].fill = light_fill
ws_val.merge_cells(f'A{row}:D{row}')

ws_val['A20'] = "EBE moyen prévisionnel"
ws_val['B20'] = "=AVERAGE(Calculs_DCF!C6:G6)"
ws_val['B20'].number_format = '#,##0'
ws_val['C20'] = "€"

ws_val['A21'] = "Multiple EBE"
ws_val['B21'] = "=Hypotheses!B20"
ws_val['B21'].number_format = '0.0'
ws_val['C21'] = "x"

ws_val['A22'] = "Dette maximale"
ws_val['B22'] = "=B20*B21"
ws_val['B22'].number_format = '#,##0'
ws_val['C22'] = "€"
ws_val['D22'] = "Formule: EBE_moy × Multiple"

ws_val['A23'] = "APPORT EN CAPITAL MINIMUM"
ws_val['A23'].font = Font(bold=True, size=12, color="C00000")
ws_val['B23'] = "=MAX(0,B17-B22)"
ws_val['B23'].number_format = '#,##0'
ws_val['B23'].font = Font(bold=True, size=12, color="C00000")
ws_val['C23'] = "€"
ws_val['D23'] = "Formule: MAX(0, Val_titres - Dette_max)"

# Appliquer bordures
for row in range(4, 24):
    for col in ['A', 'B', 'C', 'D']:
        ws_val[f'{col}{row}'].border = thin_border

# Ajuster largeur des colonnes
ws_val.column_dimensions['A'].width = 35
ws_val.column_dimensions['B'].width = 20
ws_val.column_dimensions['C'].width = 8
ws_val.column_dimensions['D'].width = 35

# ==============================================================================
# FEUILLE 4 : BUDGET TRESORERIE
# ==============================================================================
ws_tres = wb.create_sheet("Budget_Tresorerie")

# Titre
ws_tres.merge_cells('A1:G1')
ws_tres['A1'] = "BUDGET DE TRÉSORERIE DE LA HOLDING"
ws_tres['A1'].font = title_font
ws_tres['A1'].fill = title_fill
ws_tres['A1'].alignment = Alignment(horizontal='center')

# Paramètres dette
row = 3
ws_tres[f'A{row}'] = "PARAMÈTRES DE LA DETTE SENIOR"
ws_tres[f'A{row}'].font = header_font
ws_tres[f'A{row}'].fill = light_fill
ws_tres.merge_cells(f'A{row}:D{row}')

ws_tres['A4'] = "Dette Senior contractée"
ws_tres['B4'] = "=MIN(Valorisation!B22,Valorisation!B17)"
ws_tres['B4'].number_format = '#,##0'
ws_tres['C4'] = "€"
ws_tres['D4'] = "Formule: MIN(Dette_max, Val_titres)"

ws_tres['A5'] = "Taux d'intérêt annuel"
ws_tres['B5'] = "=Hypotheses!B21"
ws_tres['B5'].number_format = '0.00%'

ws_tres['A6'] = "Durée de remboursement"
ws_tres['B6'] = "=Hypotheses!B18"
ws_tres['C6'] = "ans"

ws_tres['A7'] = "Remboursement annuel du principal"
ws_tres['B7'] = "=B4/B6"
ws_tres['B7'].number_format = '#,##0'
ws_tres['C7'] = "€"
ws_tres['D7'] = "Formule: Dette / Durée (linéaire)"

# En-têtes tableau budget
row = 9
headers_tres = ["Élément", "N+1", "N+2", "N+3", "N+4", "N+5"]
for col, header in enumerate(headers_tres, 1):
    cell = ws_tres.cell(row=row, column=col, value=header)
    cell.font = header_font
    cell.fill = light_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')

# Données du budget avec formules
budget_rows = [
    ("Trésorerie début de période", 
     "=Valorisation!B23",  # Apport en capital
     "=B14",  # Tréso fin N+1
     "=C14",  # Tréso fin N+2
     "=D14",  # etc.
     "=E14"),
    
    ("+ FCF remonté de la cible",
     "=Calculs_DCF!C15",
     "=Calculs_DCF!D15",
     "=Calculs_DCF!E15",
     "=Calculs_DCF!F15",
     "=Calculs_DCF!G15"),
    
    ("- Intérêts dette senior",
     "=$B$4*$B$5",  # Dette initiale × taux pour simplifier (ou dette restante)
     "=($B$4-$B$7)*$B$5",
     "=($B$4-2*$B$7)*$B$5",
     "=($B$4-3*$B$7)*$B$5",
     "=($B$4-4*$B$7)*$B$5"),
    
    ("- Remboursement principal",
     "=$B$7",
     "=$B$7",
     "=$B$7",
     "=$B$7",
     "=$B$7"),
    
    ("= Trésorerie fin de période",
     "=B10+B11-B12-B13",
     "=C10+C11-C12-C13",
     "=D10+D11-D12-D13",
     "=E10+E11-E12-E13",
     "=F10+F11-F12-F13"),
]

for row_idx, row_data in enumerate(budget_rows, 10):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_tres.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.number_format = '#,##0'
        if "Trésorerie fin" in row_data[0]:
            cell.font = Font(bold=True)

# Dette restante
row = 17
ws_tres[f'A{row}'] = "DETTE RESTANTE"
ws_tres[f'A{row}'].font = header_font
ws_tres[f'A{row}'].fill = light_fill
ws_tres.merge_cells(f'A{row}:F{row}')

ws_tres['A18'] = "Dette restante fin de période"
ws_tres['B18'] = "=$B$4-$B$7"
ws_tres['C18'] = "=$B$4-2*$B$7"
ws_tres['D18'] = "=$B$4-3*$B$7"
ws_tres['E18'] = "=$B$4-4*$B$7"
ws_tres['F18'] = "=$B$4-5*$B$7"
for col in range(1, 7):
    ws_tres.cell(row=18, column=col).border = thin_border
    ws_tres.cell(row=18, column=col).number_format = '#,##0'

# Ajuster largeur des colonnes
ws_tres.column_dimensions['A'].width = 35
for col in range(2, 7):
    ws_tres.column_dimensions[get_column_letter(col)].width = 15

# ==============================================================================
# FEUILLE 5 : RATIOS ANALYSE FINANCIERE
# ==============================================================================
ws_ratios = wb.create_sheet("Ratios_Analyse")

# Titre
ws_ratios.merge_cells('A1:E1')
ws_ratios['A1'] = "RATIOS D'ANALYSE FINANCIÈRE"
ws_ratios['A1'].font = title_font
ws_ratios['A1'].fill = title_fill
ws_ratios['A1'].alignment = Alignment(horizontal='center')

# Explication
ws_ratios['A3'] = "Ces ratios doivent être calculés à partir des comptes annuels réels de Bricorama"
ws_ratios['A3'].font = Font(italic=True)
ws_ratios.merge_cells('A3:E3')

# En-têtes
row = 5
headers_ratios = ["Ratio", "Formule", "2022", "2023", "2024"]
for col, header in enumerate(headers_ratios, 1):
    cell = ws_ratios.cell(row=row, column=col, value=header)
    cell.font = header_font
    cell.fill = light_fill
    cell.border = thin_border

ratios_data = [
    ("Liquidité Générale", "Actif Circulant / Passif Circulant", "", "", ""),
    ("Liquidité Réduite", "(Actif Circ. - Stocks) / Passif Circ.", "", "", ""),
    ("Autonomie Financière", "Capitaux Propres / Total Passif", "", "", ""),
    ("Ratio Dettes/EBE", "Dettes Financières / EBE", "", "", ""),
    ("Rentabilité économique (ROA)", "EBIT / Total Actif", "", "", ""),
    ("Rentabilité financière (ROE)", "Résultat Net / Capitaux Propres", "", "", ""),
    ("Marge nette", "Résultat Net / CA", "", "", ""),
    ("Rotation des actifs", "CA / Total Actif", "", "", ""),
]

for row_idx, row_data in enumerate(ratios_data, 6):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_ratios.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border

# Ajuster largeur des colonnes
ws_ratios.column_dimensions['A'].width = 30
ws_ratios.column_dimensions['B'].width = 40
ws_ratios.column_dimensions['C'].width = 12
ws_ratios.column_dimensions['D'].width = 12
ws_ratios.column_dimensions['E'].width = 12

# ==============================================================================
# SAUVEGARDER LE FICHIER
# ==============================================================================
filename = "Bricorama_DCF_Acquisition.xlsx"
wb.save(filename)
print(f"✓ Fichier Excel créé avec succès : {filename}")
print("\nLe fichier contient les feuilles suivantes :")
print("  1. Hypotheses - Tous les paramètres de base et taux")
print("  2. Calculs_DCF - Projection des flux sur 5 ans avec formules")
print("  3. Valorisation - VAN, Valeur Terminale, Valeur d'Entreprise")
print("  4. Budget_Tresorerie - Budget de la Holding avec remboursement dette")
print("  5. Ratios_Analyse - Template pour les ratios d'analyse financière")
print("\n⚠️  Toutes les cellules de calcul contiennent des FORMULES Excel,")
print("    pas des valeurs en dur. Modifiez les hypothèses pour voir les résultats changer.")
