#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Analyse financière de Bricorama France - Valorisation DCF et structuration LBO
Devoir Maison - Financement et valorisation
"""

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

plt.rcParams['font.family'] = 'DejaVu Sans'
plt.rcParams['axes.unicode_minus'] = False

# =============================================================================
# 1. DONNÉES HISTORIQUES BRICORAMA (k EUR) - Source: Pappers
# =============================================================================

data_hist = {
    'annee': [2022, 2023, 2024],
    'ca': [165420, 175280, 185640],
    'ebe': [5200, 5600, 6100],
    'dotations': [1800, 1900, 2100],
    'stocks': [28500, 30100, 32400],
    'creances_clients': [8200, 9100, 9800],
    'dettes_fournisseurs': [18500, 20300, 21500],
    'dettes_financieres': [12400, 11800, 10900],
    'disponibilites': [2500, 2800, 3200],
    'capitaux_propres': [29300, 31200, 33100],
    'capital_social': 7500,
    'actif_immobilise': 26100,
}

df = pd.DataFrame(data_hist).set_index('annee')

# Calculs derivados
df['actif_circulant'] = df['stocks'] + df['creances_clients'] + df['disponibilites']  # simplificado
df['passif_circulant'] = df['dettes_fournisseurs'] + (df['dettes_financieres'] * 0.2)  # parte corrente
df['total_bilan'] = df['actif_immobilise'] + df['actif_circulant']

print("=" * 60)
print("BRICORAMA FRANCE - ANALYSE FINANCIERE")
print("=" * 60)

# =============================================================================
# 2. RATIOS HISTORIQUES
# =============================================================================

df['liquidite_generale'] = df['actif_circulant'] / df['passif_circulant']
df['liquidite_reduite'] = (df['actif_circulant'] - df['stocks']) / df['passif_circulant']
df['taux_endettement'] = df['dettes_financieres'] / df['capitaux_propres']
df['capacite_remboursement'] = df['dettes_financieres'] / df['ebe']
df['autonomie_financiere'] = df['capitaux_propres'] / df['total_bilan']
df['ratio_legal'] = df['capitaux_propres'] / df['capital_social']
df['marge_ebe_pct'] = df['ebe'] / df['ca'] * 100  # colonne nommee differemment pour eviter conflit

print("\n=== RATIOS DE LIQUIDITE ===")
print(f"2022: LG={df.loc[2022, 'liquidite_generale']:.2f}, LR={df.loc[2022, 'liquidite_reduite']:.2f}")
print(f"2023: LG={df.loc[2023, 'liquidite_generale']:.2f}, LR={df.loc[2023, 'liquidite_reduite']:.2f}")
print(f"2024: LG={df.loc[2024, 'liquidite_generale']:.2f}, LR={df.loc[2024, 'liquidite_reduite']:.2f}")

print("\n=== RATIOS DE VULNERABILITE ===")
print(f"2022: Taux endett={df.loc[2022, 'taux_endettement']:.2f}, Cap remb={df.loc[2022, 'capacite_remboursement']:.2f}ans")
print(f"2023: Taux endett={df.loc[2023, 'taux_endettement']:.2f}, Cap remb={df.loc[2023, 'capacite_remboursement']:.2f}ans")
print(f"2024: Taux endett={df.loc[2024, 'taux_endettement']:.2f}, Cap remb={df.loc[2024, 'capacite_remboursement']:.2f}ans")

# =============================================================================
# 3. HYPOTHESES DCF
# =============================================================================

print("\n" + "=" * 60)
print("HYPOTHESES DCF")
print("=" * 60)

# Paramètres
croissance_volume = 0.05
inflation = 0.05
croissance_nominale = (1 + croissance_volume) * (1 + inflation) - 1  # 10.25%
print(f"Croissance nominale: {croissance_nominale:.2%}")

marge_ebe = float(df.loc[2024, 'ebe']) / float(df.loc[2024, 'ca'])  # ~3.29%
print(f"Marge EBE 2024: {marge_ebe*100:.2%}")

taux_dotation = df['dotations'].iloc[-1] / df['ca'].iloc[-1]
print(f"Taux dotations/CA: {taux_dotation*100:.2%}")

taux_invest = 0.02
taux_is = 0.25
wacc = 0.08
g = 0.02
multiple_ebe = 3.5
taux_dette = 0.05
horizon = 5

# BFR normatif (jours de CA)
ca_moyen = df.loc[2023:2024, 'ca'].mean()
stocks_moyen = df.loc[2023:2024, 'stocks'].mean()
clients_moyen = df.loc[2023:2024, 'creances_clients'].mean()
fourn_moyen = df.loc[2023:2024, 'dettes_fournisseurs'].mean()
achats_moyen = (104500 + 110800) / 2

jours_stocks = (stocks_moyen / ca_moyen) * 360
jours_clients = (clients_moyen / ca_moyen) * 360
jours_fourn = (fourn_moyen / achats_moyen) * 360
bfr_jours = jours_stocks + jours_clients - jours_fourn
coef_bfr = bfr_jours / 360

print(f"\nBFR normatif: {bfr_jours:.1f} jours ({coef_bfr*100:.2f}% CA)")

# =============================================================================
# 4. PROJECTION DCF 2025-2029
# =============================================================================

print("\n" + "=" * 60)
print("TABLEAU DCF (k EUR)")
print("=" * 60)

ca_base = df.loc[2024, 'ca']
annees = [2025, 2026, 2027, 2028, 2029]

# Projections
ca_list = [ca_base * (1 + croissance_nominale)**i for i in range(1, horizon+1)]
ebe_list = [ca * marge_ebe for ca in ca_list]
dot_list = [ca * taux_dotation for ca in ca_list]
ebit_list = [ebe - dot_ for ebe, dot_ in zip(ebe_list, dot_list)]
impot_list = [ebit_ * taux_is for ebit_ in ebit_list]
invest_list = [ca * taux_invest for ca in ca_list]

# BFR
bfr_initial = ca_base * coef_bfr
bfr_list = [ca * coef_bfr for ca in ca_list]
var_bfr = [bfr_list[0] - bfr_initial] + [bfr_list[i] - bfr_list[i-1] for i in range(1, horizon)]

# FCF CORRIGE: FCF = EBE - Impot(EBIT) - Investissements - Variation BFR
fcf_list = [ebe - imp - inv - var_ for ebe, imp, inv, var_ in zip(ebe_list, impot_list, invest_list, var_bfr)]

# Actualisation
coeffs = [1 / (1 + wacc)**(i+1) for i in range(horizon)]
fcf_act = [fcf * coeff for fcf, coeff in zip(fcf_list, coeffs)]

# Affichage
print(f"{'2024':>6} | {'CA':>10} | {'EBE':>8} | {'EBIT':>8} | {'Impôt':>8} | {'FCF':>10}")
print("-" * 65)
print(f"{'Base':>6} | {ca_base:>10,.0f} | {df.loc[2024,'ebe']:>8,.0f} | {df.loc[2024,'ebe']-df.loc[2024,'dotations']:>8,.0f} | {(df.loc[2024,'ebe']-df.loc[2024,'dotations'])*0.25:>8,.0f} |")
for i, (ca, ebe, ebit, imp, fcf) in enumerate(zip(ca_list, ebe_list, ebit_list, impot_list, fcf_list)):
    print(f"{annees[i]:>6} | {ca:>10,.0f} | {ebe:>8,.0f} | {ebit:>8,.0f} | {imp:>8,.0f} | {fcf:>10,.0f}")

print(f"\nSomme FCF actualisés: {sum(fcf_act):,.0f} k EUR")

# =============================================================================
# 5. VALORISATION
# =============================================================================

print("\n" + "=" * 60)
print("VALORISATION")
print("=" * 60)

# Valeur terminale (Gordon-Shapiro)
vt = fcf_list[-1] * (1 + g) / (wacc - g)
vt_act = vt * coeffs[-1]

ve = sum(fcf_act) + vt_act
dette_nette = df.loc[2024, 'dettes_financieres'] - df.loc[2024, 'disponibilites']
valeur_titres = ve - dette_nette

print(f"FCF actualisés (5 ans): {sum(fcf_act):,.0f} k EUR")
print(f"Valeur terminale: {vt:,.0f} k EUR")
print(f"VT actualisée: {vt_act:,.0f} k EUR")
print(f"\n>>> VALEUR D'ENTREPRISE (VE): {ve:,.0f} k EUR ({ve/1000:.1f} M EUR)")
print(f"Dette nette 2024: {dette_nette:,.0f} k EUR")
print(f">>> VALEUR DES TITRES: {valeur_titres:,.0f} k EUR ({valeur_titres/1000:.1f} M EUR)")

# =============================================================================
# 6. STRUCTURATION FINANCIERE
# =============================================================================

print("\n" + "=" * 60)
print("STRUCTURATION DU FINANCEMENT")
print("=" * 60)

ebe_moyen = np.mean(ebe_list)
dette_max = multiple_ebe * ebe_moyen

print(f"EBE moyen: {ebe_moyen:,.0f} k EUR")
print(f"Dette max ({multiple_ebe}x EBE): {dette_max:,.0f} k EUR")

# Cas 1: dette non consolidée
print("\n--- CAS 1: DETTE NON CONSOLIDEE ---")
dette1 = min(dette_max, valeur_titres)
apport1 = valeur_titres - dette1
print(f"Dette senior holding: {dette1:,.0f} k EUR")
print(f"APPORT EN CAPITAL: {apport1:,.0f} k EUR ({apport1/1000:.1f} M EUR)")

# Cas 2: dette consolidée
print("\n--- CAS 2: DETTE CONSOLIDEE ---")
dette_existante = df.loc[2024, 'dettes_financieres']
dette2 = max(0, dette_max - dette_existante)
apport2 = valeur_titres - dette2
print(f"Dette existante cible: {dette_existante:,.0f} k EUR")
print(f"Dette holding max: {dette2:,.0f} k EUR")
print(f"APPORT EN CAPITAL: {apport2:,.0f} k EUR ({apport2/1000:.1f} M EUR)")

# =============================================================================
# 7. BUDGET DE TRESORERIE HOLDING (Cas 1)
# =============================================================================

print("\n" + "=" * 60)
print("BUDGET TRESORERIE HOLDING (k EUR)")
print("=" * 60)

remboursement = dette1 / horizon
tresor = apport1
dette_rest = dette1

print(f"{'Annee':>6} | {'TresDebut':>10} | {'FCF':>8} | {'Interets':>8} | {'Rembours':>8} | {'TresFin':>10} | {'Dette':>10}")
print("-" * 75)

budget = []
for i, annee in enumerate(annees):
    interets = dette_rest * taux_dette
    tresor_fin = tresor + fcf_list[i] - interets - remboursement
    budget.append({'annee': annee, 'tres_debut': tresor, 'fcf': fcf_list[i], 
                   'interets': interets, 'remboursement': remboursement, 
                   'tres_fin': tresor_fin, 'dette': dette_rest - remboursement})
    print(f"{annee:>6} | {tresor:>10,.0f} | {fcf_list[i]:>8,.0f} | {interets:>8,.0f} | {remboursement:>8,.0f} | {tresor_fin:>10,.0f} | {dette_rest-remboursement:>10,.0f}")
    tresor = tresor_fin
    dette_rest -= remboursement

# =============================================================================
# 8. GRAPHIQUES
# =============================================================================

print("\n" + "=" * 60)
print("GENERATION GRAPHIQUES")
print("=" * 60)

colors = {'blue': '#4472C4', 'orange': '#ED7D31', 'green': '#70AD47'}

# Graphique 1: CA et EBE historique
fig, axes = plt.subplots(2, 3, figsize=(15, 10))

axes[0,0].bar([2022,2023,2024], df['ca']/1000, color=colors['blue'], alpha=0.8)
axes[0,0].set_title('Chiffre Affaire (M EUR)')
axes[0,0].set_ylabel('M EUR')

axes[0,1].bar([2022,2023,2024], df['ebe']/1000, color=colors['orange'], alpha=0.8)
axes[0,1].set_title('EBE (M EUR)')

axes[0,2].plot([2022,2023,2024], df['liquidite_generale'], 'o-', color=colors['blue'], linewidth=2)
axes[0,2].plot([2022,2023,2024], df['liquidite_reduite'], 's-', color=colors['orange'], linewidth=2)
axes[0,2].axhline(y=1, color='green', linestyle='--', label='Seuil LG>1')
axes[0,2].set_title('Ratios Liquidité')
axes[0,2].legend()

axes[1,0].bar([2022,2023,2024], df['capacite_remboursement'], color=colors['green'], alpha=0.8)
axes[1,0].axhline(y=4, color='red', linestyle='--', label='Alerte >4 ans')
axes[1,0].set_title('Capacité Remboursement (ans)')
axes[1,0].legend()

axes[1,1].bar(annees, [f/1000 for f in fcf_list], color=colors['blue'], alpha=0.8, label='FCF')
axes[1,1].plot(annees, [f/1000 for f in fcf_act], 'o-', color=colors['orange'], linewidth=2, label='FCF actu.')
axes[1,1].set_title('Projection FCF (M EUR)')
axes[1,1].legend()

# Graphique décomposition valeur
labels = ['FCF actu.', 'VT actu.', 'VE', 'Dette', 'Titres']
values = [sum(fcf_act)/1000, vt_act/1000, ve/1000, -dette_nette/1000, valeur_titres/1000]
pos = [0, 1, 2.5, 3.5, 4.5]
axes[1,2].bar(pos[:2], values[:2], color=[colors['blue'], colors['orange']])
axes[1,2].bar(2, values[2], color=colors['green'])
axes[1,2].bar(3, values[3], color='red')
axes[1,2].bar(4, values[4], color=colors['green'])
axes[1,2].set_xticks(pos)
axes[1,2].set_xticklabels(labels, fontsize=8)
axes[1,2].set_title('Décomposition Valeur (M EUR)')
axes[1,2].axhline(y=0, color='black')

plt.tight_layout()
plt.savefig('bricorama_graphiques.png', dpi=150, bbox_inches='tight')
print("- bricorama_graphiques.png")
plt.close()

# =============================================================================
# 9. GENERATION EXCEL
# =============================================================================

print("\n" + "=" * 60)
print("GENERATION EXCEL")
print("=" * 60)

wb = Workbook()
ws = wb.active
ws.title = "Resume"

ws['A1'] = "BRICORAMA FRANCE - ANALYSE FINANCIERE"
ws['A1'].font = Font(bold=True, size=14)

# Hypotheses
ws['A3'] = "HYPOTHESES"
ws['A3'].font = Font(bold=True)
data_hyp = [
    ("Croissance nominale", f"{croissance_nominale*100:.2f}%"),
    ("Marge EBE", f"{marge_ebe*100:.2f}%"),
    ("Taux dotations", f"{taux_dotation*100:.2f}%"),
    ("Taux investissement", f"{taux_invest*100:.0f}%"),
    ("Coefficient BFR", f"{coef_bfr*100:.2f}%"),
    ("WACC", f"{wacc*100:.0f}%"),
    ("g (croissance perpétuelle)", f"{g*100:.0f}%"),
    ("Taux IS", f"{taux_is*100:.0f}%"),
    ("Multiple EBE dette", multiple_ebe),
]
for i, (k, v) in enumerate(data_hyp, start=4):
    ws[f'A{i}'] = k
    ws[f'B{i}'] = v

# Resultats valorisation
ws['A14'] = "VALORISATION"
ws['A14'].font = Font(bold=True)
ws['A15'] = "Somme FCF actualisés (5 ans)"
ws['B15'] = sum(fcf_act)
ws['A16'] = "Valeur Terminale actualisée"
ws['B16'] = vt_act
ws['A17'] = "VALEUR D'ENTREPRISE"
ws['B17'] = ve
ws['A18'] = "Dette nette"
ws['B18'] = dette_nette
ws['A19'] = "VALEUR DES TITRES"
ws['B19'] = valeur_titres
ws['B17'].font = Font(bold=True)
ws['B19'].font = Font(bold=True)

# Financement
ws['A22'] = "FINANCEMENT"
ws['A22'].font = Font(bold=True)
ws['A23'] = "Cas 1 - Dette non consolidée"
ws['B23'] = apport1
ws['A24'] = "Cas 2 - Dette consolidée"
ws['B24'] = apport2

wb.save('bricorama_analyse.xlsx')
print("- bricorama_analyse.xlsx")

# =============================================================================
# 10. RESUME FINAL
# =============================================================================

print("\n" + "=" * 60)
print("RESUME FINAL")
print("=" * 60)
print(f"""
VALORISATION DCF - BRICORAMA FRANCE
===================================
Valeur d'Entreprise (VE): {ve:,.0f} k EUR ({ve/1000:.1f} M EUR)
Dette nette 2024: {dette_nette:,.0f} k EUR
VALEUR DES TITRES: {valeur_titres:,.0f} k EUR ({valeur_titres/1000:.1f} M EUR)

STRUCTURATION LBO:
- Cas 1 (non consolidé): Apport = {apport1:,.0f} k EUR ({apport1/1000:.1f} M EUR)
- Cas 2 (consolidé):     Apport = {apport2:,.0f} k EUR ({apport2/1000:.1f} M EUR)

FICHIERS GENERES:
- bricorama_graphiques.png
- bricorama_analyse.xlsx
""")

print("=== FIN DU SCRIPT ===")