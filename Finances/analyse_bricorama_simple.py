#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Bricorama France - Analyse financiere simplifiee"""

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side

plt.rcParams['font.family'] = 'DejaVu Sans'
print("=" * 60)
print("BRICORAMA FRANCE - ANALYSE FINANCIERE")
print("=" * 60)

# =============================================================================
# 1. DONNEES HISTORIQUES (k EUR)
# =============================================================================

CA_2022, CA_2023, CA_2024 = 165420, 175280, 185640
EBE_2022, EBE_2023, EBE_2024 = 5200, 5600, 6100
Dette_2022, Dette_2023, Dette_2024 = 12400, 11800, 10900
Tresor_2022, Tresor_2023, Tresor_2024 = 2500, 2800, 3200
CP_2022, CP_2023, CP_2024 = 29300, 31200, 33100

Dette_nette_2024 = Dette_2024 - Tresor_2024  # 7700 k EUR

print(f"\nDONNEES 2024:")
print(f"  CA: {CA_2024:,} k EUR")
print(f"  EBE: {EBE_2024:,} k EUR")
print(f"  Dettes fin.: {Dette_2024:,} k EUR")
print(f"  Trésorerie: {Tresor_2024:,} k EUR")
print(f"  Dette nette: {Dette_nette_2024:,} k EUR")

# =============================================================================
# 2. RATIOS HISTORIQUES
# =============================================================================

Liquidite_2022 = 1.71  # Donnees reelles calculees
Liquidite_2023 = 1.68
Liquidite_2024 = 1.62

Cap_remb_2022 = Dette_2022 / EBE_2022  # 2.38
Cap_remb_2023 = Dette_2023 / EBE_2023  # 2.11
Cap_remb_2024 = Dette_2024 / EBE_2024  # 1.79

print(f"\nRATIOS CLES:")
print(f"  Capacité remboursement 2024: {Cap_remb_2024:.2f} ans")
print(f"  Taux endettement 2024: {Dette_2024/CP_2024:.2f}")

# =============================================================================
# 3. HYPOTHESES DCF
# =============================================================================

croissance = 0.1025  # 10.25% (5% vol + 5% prix)
marge_ebe = 0.033  # 3.30% (moyenne historique)
taux_dotation = 0.011  # 1.1% du CA
taux_invest = 0.02  # 2% du CA
taux_is = 0.25
wacc = 0.08
g = 0.02
multiple_ebe = 3.5
taux_dette = 0.05

print(f"\nHYPOTHESES:")
print(f"  Croissance CA: {croissance*100:.2f}%")
print(f"  Marge EBE: {marge_ebe*100:.2f}%")
print(f"  WACC: {wacc*100:.0f}%")

# =============================================================================
# 4. PROJECTION DCF 2025-2029
# =============================================================================

print(f"\n" + "=" * 60)
print("TABLEAU DCF (k EUR)")
print("=" * 60)

annees = [2025, 2026, 2027, 2028, 2029]
ca_list = []
ebe_list = []
dot_list = []
ebit_list = []
impot_list = []
invest_list = []
bfr_list = []
var_bfr_list = []
fcf_list = []

CA = CA_2024
BFR_initial = CA_2024 * 0.0314  # 3.14% du CA

print(f"{'Année':>6} | {'CA':>10} | {'EBE':>8} | {'EBIT':>8} | {'FCF':>10}")
print("-" * 55)

for i, annee in enumerate(annees):
    CA = CA * (1 + croissance) if i > 0 else CA_2024 * (1 + croissance)
    EBE = CA * marge_ebe
    Dot = CA * taux_dotation
    EBIT = EBE - Dot
    Impot = EBIT * taux_is
    Invest = CA * taux_invest
    
    BFR = CA * 0.0314
    Var_BFR = BFR - BFR_initial if i == 0 else BFR - (ca_list[-1] * 0.0314)
    
    FCF = EBE - Impot - Invest - Var_BFR
    
    ca_list.append(CA)
    ebe_list.append(EBE)
    dot_list.append(Dot)
    ebit_list.append(EBIT)
    impot_list.append(Impot)
    invest_list.append(Invest)
    bfr_list.append(BFR)
    var_bfr_list.append(Var_BFR)
    fcf_list.append(FCF)
    BFR_initial = BFR
    
    print(f"{annee:>6} | {CA:>10,.0f} | {EBIT:>8,.0f} | {EBIT:>8,.0f} | {FCF:>10,.0f}")

# Actualisation
coeffs = [1 / (1 + wacc) ** (i+1) for i in range(5)]
fcf_act = [f * c for f, c in zip(fcf_list, coeffs)]
somme_fcf_act = sum(fcf_act)

print(f"\nSomme FCF actualisés: {somme_fcf_act:,.0f} k EUR")

# =============================================================================
# 5. VALORISATION
# =============================================================================

VT = fcf_list[-1] * (1 + g) / (wacc - g)
VT_act = VT * coeffs[-1]
VE = somme_fcf_act + VT_act
Valeur_titres = VE - Dette_nette_2024

print(f"\n" + "=" * 60)
print("VALORISATION")
print("=" * 60)
print(f"VE (valeur d'entreprise): {VE:,.0f} k EUR ({VE/1000:.1f} M EUR)")
print(f"Dette nette: {Dette_nette_2024:,.0f} k EUR")
print(f"Valeur des titres: {Valeur_titres:,.0f} k EUR ({Valeur_titres/1000:.1f} M EUR)")

# =============================================================================
# 6. STRUCTURATION LBO
# =============================================================================

EBE_moyen = np.mean(ebe_list)
Dette_max = multiple_ebe * EBE_moyen

print(f"\n" + "=" * 60)
print("STRUCTURATION DU FINANCEMENT")
print("=" * 60)
print(f"EBE moyen: {EBE_moyen:,.0f} k EUR")
print(f"Dette max ({multiple_ebe}x EBE): {Dette_max:,.0f} k EUR")

# Cas 1
Dette1 = min(Dette_max, Valeur_titres)
Apport1 = max(0, Valeur_titres - Dette1)
print(f"\nCAS 1 - Dette non consolidee:")
print(f"  Dette holding: {Dette1:,.0f} k EUR")
print(f"  APPORT: {Apport1:,.0f} k EUR ({Apport1/1000:.1f} M EUR)")

# Cas 2
Dette_existante = Dette_2024
Dette2 = max(0, Dette_max - Dette_existante)
Apport2 = max(0, Valeur_titres - Dette2)
print(f"\nCAS 2 - Dette consolidee:")
print(f"  Dette existante: {Dette_existante:,.0f} k EUR")
print(f"  Dette holding max: {Dette2:,.0f} k EUR")
print(f"  APPORT: {Apport2:,.0f} k EUR ({Apport2/1000:.1f} M EUR)")

# =============================================================================
# 7. BUDGET TRESORERIE (Cas 1)
# =============================================================================

print(f"\n" + "=" * 60)
print("BUDGET TRESORERIE HOLDING - CAS 1")
print("=" * 60)

Remb = Dette1 / 5
Tres = Apport1
Dette = Dette1

print(f"{'Annee':>6} | {'Tres.Deb':>10} | {'FCF':>8} | {'Interet':>8} | {'Remb.':>8} | {'Tres.Fin':>10}")
print("-" * 65)

for i, annee in enumerate(annees):
    Interet = Dette * taux_dette
    Tres_fin = Tres + fcf_list[i] - Interet - Remb
    print(f"{annee:>6} | {Tres:>10,.0f} | {fcf_list[i]:>8,.0f} | {Interet:>8,.0f} | {Remb:>8,.0f} | {Tres_fin:>10,.0f}")
    Tres = Tres_fin
    Dette -= Remb

# =============================================================================
# 8. GRAPHIQUES
# =============================================================================

print(f"\n" + "=" * 60)
print("GENERATION GRAPHIQUES")
print("=" * 60)

fig, axes = plt.subplots(2, 3, figsize=(14, 9))
fig.suptitle('Bricorama France - Analyse Financiere DCF', fontsize=14, fontweight='bold')

# 1. CA et EBE historique
axes[0,0].bar([2022,2023,2024], [CA_2022/1000, CA_2023/1000, CA_2024/1000], color='#4472C4', alpha=0.8)
axes[0,0].set_title('CA (M EUR)')
axes[0,0].set_ylabel('M EUR')

# 2. Capacite remboursement
axes[0,1].bar([2022,2023,2024], [Cap_remb_2022, Cap_remb_2023, Cap_remb_2024], color='#70AD47', alpha=0.8)
axes[0,1].axhline(y=4, color='red', linestyle='--', label='Alerte >4 ans')
axes[0,1].set_title('Capacite Remboursement (ans)')
axes[0,1].legend()

# 3. Ratios liquidite
axes[0,2].plot([2022,2023,2024], [Liquidite_2022, Liquidite_2023, Liquidite_2024], 'o-', color='#4472C4', linewidth=2, label='LG')
axes[0,2].axhline(y=1, color='green', linestyle='--', label='Seuil LG=1')
axes[0,2].set_title('Liquidite Generale')
axes[0,2].legend()

# 4. FCF projection
axes[1,0].bar(annees, [f/1000 for f in fcf_list], color='#4472C4', alpha=0.8, label='FCF')
axes[1,0].plot(annees, [f/1000 for f in fcf_act], 'o-', color='#ED7D31', linewidth=2, label='FCF actu.')
axes[1,0].set_title('Projection FCF (M EUR)')
axes[1,0].legend()

# 5. Decomposition valeur
labels = ['FCF actu.', 'VT actu.', 'VE', 'Dette', 'Titres']
values = [somme_fcf_act/1000, VT_act/1000, VE/1000, -Dette_nette_2024/1000, Valeur_titres/1000]
pos = [0, 1, 2.5, 3.5, 4.5]
axes[1,1].bar(pos[:2], values[:2], color=['#4472C4', '#ED7D31'])
axes[1,1].bar(2, values[2], color='#70AD47')
axes[1,1].bar(3, values[3], color='red')
axes[1,1].bar(4, values[4], color='#70AD47')
axes[1,1].set_xticks(pos)
axes[1,1].set_xticklabels(['FCF', 'VT', 'VE', 'Dette', 'Titres'], fontsize=8)
axes[1,1].set_title('Decomposition Valeur (M EUR)')
axes[1,1].axhline(y=0, color='black')

# 6. Synthese financement
fin_data = ['Cas 1\nNon cons.', 'Cas 2\nConsolide']
fin_values = [Apport1/1000, Apport2/1000]
colors_bar = ['#4472C4', '#ED7D31'] if Apport2 > Apport1 else ['#4472C4', '#70AD47']
axes[1,2].bar(fin_data, fin_values, color=colors_bar)
axes[1,2].set_title('Apport Capital (M EUR)')
for i, v in enumerate(fin_values):
    axes[1,2].text(i, v + 0.5, f'{v:.1f}M', ha='center', fontsize=9)

plt.tight_layout()
plt.savefig('bricorama_rapport.png', dpi=150, bbox_inches='tight')
print("- bricorama_rapport.png")
plt.close()

# =============================================================================
# 9. EXCEL
# =============================================================================

print(f"\n" + "=" * 60)
print("GENERATION EXCEL")
print("=" * 60)

wb = Workbook()
ws = wb.active
ws.title = "Resume"

ws['A1'] = "BRICORAMA FRANCE - ANALYSE DCF"
ws['A1'].font = Font(bold=True, size=14)

# Hypotheses
ws['A3'] = "HYPOTHESES"
ws['A3'].font = Font(bold=True)
ws['A4'] = "Croissance CA"; ws['B4'] = f"{croissance*100:.2f}%"
ws['A5'] = "Marge EBE"; ws['B5'] = f"{marge_ebe*100:.2f}%"
ws['A6'] = "WACC"; ws['B6'] = f"{wacc*100:.0f}%"
ws['A7'] = "g (perpetuite)"; ws['B7'] = f"{g*100:.0f}%"

# Resultats
ws['A9'] = "VALORISATION"
ws['A9'].font = Font(bold=True)
ws['A10'] = "VE"; ws['B10'] = VE
ws['A11'] = "Dette nette"; ws['B11'] = Dette_nette_2024
ws['A12'] = "VALEUR TITRES"; ws['B12'] = Valeur_titres
ws['B12'].font = Font(bold=True)

# Financement
ws['A14'] = "FINANCEMENT"
ws['A14'].font = Font(bold=True)
ws['A15'] = "Cas 1 (non consolide)"; ws['B15'] = Apport1
ws['A16'] = "Cas 2 (consolide)"; ws['B16'] = Apport2

# DCF
ws['A18'] = "DCF (k EUR)"
ws['A18'].font = Font(bold=True)
ws['A19'] = "Année"; ws['B19'] = "CA"; ws['C19'] = "EBE"; ws['D19'] = "FCF"; ws['E19'] = "FCF actu."
for i, annee in enumerate(annees):
    ws.cell(row=20+i, column=1, value=annee)
    ws.cell(row=20+i, column=2, value=round(ca_list[i]))
    ws.cell(row=20+i, column=3, value=round(ebe_list[i]))
    ws.cell(row=20+i, column=4, value=round(fcf_list[i]))
    ws.cell(row=20+i, column=5, value=round(fcf_act[i]))

wb.save('bricorama_rapport.xlsx')
print("- bricorama_rapport.xlsx")

# =============================================================================
# 10. RESUME FINAL
# =============================================================================

print(f"\n" + "=" * 60)
print("RESUME FINAL")
print("=" * 60)
print(f"""
VALEUR D'ENTREPRISE (VE): {VE:,.0f} k EUR ({VE/1000:.1f} M EUR)
VALEUR DES TITRES: {Valeur_titres:,.0f} k EUR ({Valeur_titres/1000:.1f} M EUR)

FINANCEMENT LBO:
  Cas 1 (non consolide): Apport = {Apport1:,.0f} k EUR
  Cas 2 (consolide):     Apport = {Apport2:,.0f} k EUR

FICHIERS GENERES:
  - bricorama_rapport.png (graphiques)
  - bricorama_rapport.xlsx (tableaux)
""")

print("=== FIN ===")