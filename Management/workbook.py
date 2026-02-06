from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter

def create_maas_porter_excel(filename="maas_porter.xlsx"):
    # -------------------------
    # Données de base
    # -------------------------
    activities = [
        ("Logistique interne", 180, {"DF": 120, "DV": 30, "IF": 20, "IV": 10}),
        ("Opérations", 260, {"DF": 170, "DV": 40, "IF": 30, "IV": 20}),
        ("Logistique externe", 120, {"DF": 70, "DV": 20, "IF": 20, "IV": 10}),
        ("Marketing & vente", 120, {"DF": 60, "DV": 30, "IF": 20, "IV": 10}),
        ("Services", 60, {"DF": 30, "DV": 10, "IF": 10, "IV": 10}),
        ("Infrastructure de l’entreprise", 140, {"DF": 60, "DV": 20, "IF": 40, "IV": 20}),
        ("Gestion des RH", 60, {"DF": 30, "DV": 10, "IF": 10, "IV": 10}),
        ("Développement technologique", 160, {"DF": 90, "DV": 30, "IF": 25, "IV": 15}),
        ("Approvisionnement", 100, {"DF": 40, "DV": 20, "IF": 25, "IV": 15}),
    ]

    # -------------------------
    # Création du classeur
    # -------------------------
    wb = Workbook()

    # Styles de base
    header_font = Font(bold=True)
    title_font = Font(bold=True, size=14)
    currency_style = NamedStyle(name="currency_style")
    currency_style.number_format = u'#,##0.00 €'
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    if "currency_style" not in wb.named_styles:
        wb.add_named_style(currency_style)

    # -------------------------
    # Feuille Hypothèses
    # -------------------------
    ws_h = wb.active
    ws_h.title = "Hypothèses"

    ws_h["A1"] = "Hypothèses générales"
    ws_h["A1"].font = title_font

    ws_h["A3"] = "Durée du projet (mois)"
    ws_h["B3"] = 6

    ws_h["A4"] = "Jours ouvrés par mois"
    ws_h["B4"] = 20

    ws_h["A5"] = "Nombre de personnes"
    ws_h["B5"] = 10

    ws_h["A6"] = "TJM (€/jour/personne)"
    ws_h["B6"] = 600

    ws_h["A8"] = "Total personnes-jour (calculé)"
    ws_h["B8"] = "=B3*B4*B5"  # 6 * 20 * 10 = 1200

    # Tableau de répartition des personnes-jour par activité
    ws_h["A10"] = "Répartition des personnes-jour par activité"
    ws_h["A10"].font = header_font

    ws_h["A12"] = "Activité"
    ws_h["B12"] = "Total pj"
    ws_h["C12"] = "DF pj"
    ws_h["D12"] = "DV pj"
    ws_h["E12"] = "IF pj"
    ws_h["F12"] = "IV pj"
    for col in range(1, 7):
        ws_h.cell(row=12, column=col).font = header_font
        ws_h.cell(row=12, column=col).border = thin_border
        ws_h.cell(row=12, column=col).alignment = Alignment(horizontal="center")

    start_row = 13
    for i, (name, total_pj, breakdown) in enumerate(activities):
        r = start_row + i
        ws_h[f"A{r}"] = name
        ws_h[f"B{r}"] = total_pj
        ws_h[f"C{r}"] = breakdown["DF"]
        ws_h[f"D{r}"] = breakdown["DV"]
        ws_h[f"E{r}"] = breakdown["IF"]
        ws_h[f"F{r}"] = breakdown["IV"]
        for col in range(1, 7):
            ws_h.cell(row=r, column=col).border = thin_border

    # Total ligne
    total_row = start_row + len(activities)
    ws_h[f"A{total_row}"] = "Total"
    ws_h[f"A{total_row}"].font = header_font
    for col in range(2, 7):
        col_letter = get_column_letter(col)
        ws_h[f"{col_letter}{total_row}"] = f"=SUM({col_letter}{start_row}:{col_letter}{total_row-1})"
        ws_h[f"{col_letter}{total_row}"].font = header_font
        ws_h[f"{col_letter}{total_row}"].border = thin_border

    # Ajustement largeur colonnes
    for col in range(1, 7):
        ws_h.column_dimensions[get_column_letter(col)].width = 25

    # -------------------------
    # Feuilles par activité
    # -------------------------
    activity_sheets = {}  # pour la synthèse

    for idx, (name, total_pj, breakdown) in enumerate(activities, start=1):
        ws = wb.create_sheet(title=f"{idx}. {name[:25]}")  # limiter la longueur du nom

        # Titre
        ws["A1"] = f"Activité {idx} : {name}"
        ws["A1"].font = title_font

        # Rappel des hypothèses (liens vers Hypothèses)
        ws["A3"] = "TJM (€/jour/personne)"
        ws["B3"] = "=Hypothèses!B6"

        ws["A4"] = "Personnes-jour DF"
        ws["B4"] = f"=Hypothèses!C{start_row + idx - 1}"

        ws["A5"] = "Personnes-jour DV"
        ws["B5"] = f"=Hypothèses!D{start_row + idx - 1}"

        ws["A6"] = "Personnes-jour IF"
        ws["B6"] = f"=Hypothèses!E{start_row + idx - 1}"

        ws["A7"] = "Personnes-jour IV"
        ws["B7"] = f"=Hypothèses!F{start_row + idx - 1}"

        for row in range(3, 8):
            ws[f"A{row}"].font = header_font

        # Tableau de calcul
        ws["A9"] = "Type de coût"
        ws["B9"] = "Personnes-jour"
        ws["C9"] = "TJM"
        ws["D9"] = "Coût (€)"
        for col in range(1, 5):
            ws.cell(row=9, column=col).font = header_font
            ws.cell(row=9, column=col).border = thin_border
            ws.cell(row=9, column=col).alignment = Alignment(horizontal="center")

        rows = {
            "DF": 10,
            "DV": 11,
            "IF": 12,
            "IV": 13,
        }

        ws["A10"] = "DF (direct fixe)"
        ws["A11"] = "DV (direct variable)"
        ws["A12"] = "IF (indirect fixe)"
        ws["A13"] = "IV (indirect variable)"

        # Lien personnes-jour
        ws["B10"] = "=B4"
        ws["B11"] = "=B5"
        ws["B12"] = "=B6"
        ws["B13"] = "=B7"

        # TJM
        ws["C10"] = "=B3"
        ws["C11"] = "=B3"
        ws["C12"] = "=B3"
        ws["C13"] = "=B3"

        # Coût = pj * TJM
        ws["D10"] = "=B10*C10"
        ws["D11"] = "=B11*C11"
        ws["D12"] = "=B12*C12"
        ws["D13"] = "=B13*C13"

        # Formatage
        for r in range(10, 14):
            for c in range(1, 5):
                cell = ws.cell(row=r, column=c)
                cell.border = thin_border
            ws[f"D{r}"].style = "currency_style"

        # Total
        ws["C15"] = "Total activité"
        ws["C15"].font = header_font
        ws["D15"] = "=SUM(D10:D13)"
        ws["D15"].style = "currency_style"
        ws["D15"].font = Font(bold=True, underline="single")

        # Ajustement largeur colonnes
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 18

        # Sauvegarde de la référence pour la synthèse
        activity_sheets[name] = ws.title

    # -------------------------
    # Feuille Synthèse
    # -------------------------
    ws_s = wb.create_sheet(title="Synthèse")

    ws_s["A1"] = "Synthèse des coûts par activité (Porter)"
    ws_s["A1"].font = title_font

    ws_s["A3"] = "Activité de Porter"
    ws_s["B3"] = "DF (€)"
    ws_s["C3"] = "DV (€)"
    ws_s["D3"] = "IF (€)"
    ws_s["E3"] = "IV (€)"
    ws_s["F3"] = "Total (€)"

    for col in range(1, 7):
        ws_s.cell(row=3, column=col).font = header_font
        ws_s.cell(row=3, column=col).border = thin_border
        ws_s.cell(row=3, column=col).alignment = Alignment(horizontal="center")

    synth_start_row = 4
    for idx, (name, _, _) in enumerate(activities, start=1):
        r = synth_start_row + idx - 1
        sheet_name = f"{idx}. {name[:25]}"

        ws_s[f"A{r}"] = name

        # DF = D10, DV = D11, IF = D12, IV = D13, Total = D15
        ws_s[f"B{r}"] = f"='{sheet_name}'!D10"
        ws_s[f"C{r}"] = f"='{sheet_name}'!D11"
        ws_s[f"D{r}"] = f"='{sheet_name}'!D12"
        ws_s[f"E{r}"] = f"='{sheet_name}'!D13"
        ws_s[f"F{r}"] = f"='{sheet_name}'!D15"

        for col in range(2, 7):
            cell = ws_s.cell(row=r, column=col)
            cell.style = "currency_style"
            cell.border = thin_border

    # Ligne total
    total_row_s = synth_start_row + len(activities)
    ws_s[f"A{total_row_s}"] = "Total"
    ws_s[f"A{total_row_s}"].font = header_font

    for col in range(2, 7):
        col_letter = get_column_letter(col)
        ws_s[f"{col_letter}{total_row_s}"] = f"=SUM({col_letter}{synth_start_row}:{col_letter}{total_row_s-1})"
        ws_s[f"{col_letter}{total_row_s}"].style = "currency_style"
        ws_s[f"{col_letter}{total_row_s}"].font = Font(bold=True, underline="single")
        ws_s[f"{col_letter}{total_row_s}"].border = thin_border

    # Ajustement largeur colonnes
    ws_s.column_dimensions["A"].width = 35
    for col in range(2, 7):
        ws_s.column_dimensions[get_column_letter(col)].width = 18

    # -------------------------
    # Sauvegarde du fichier
    # -------------------------
    wb.save(filename)
    print(f"Fichier Excel généré : {filename}")


if __name__ == "__main__":
    create_maas_porter_excel("maas_porter.xlsx")
